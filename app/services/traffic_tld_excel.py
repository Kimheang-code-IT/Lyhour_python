"""Read Austroads Appendix TLD workbooks (axle load distribution %)."""
from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from app.services.esal_calculator import AXLE_TYPE_ORDER, has_tld_distribution

_TLD_AXLE_LABELS = frozenset(AXLE_TYPE_ORDER)


def read_tld_workbook(path: str | Path) -> dict[str, Any]:
    """Load a TLD Excel file and parse axle load distribution rows."""
    file_path = Path(path)
    if not file_path.is_file():
        raise FileNotFoundError(f"TLD file not found: {file_path}")

    suffix = file_path.suffix.lower()
    if suffix not in {".xlsx", ".xlsm", ".xls"}:
        raise ValueError("TLD file must be an Excel workbook (.xlsx, .xls, or .xlsm).")

    distribution_rows: list[dict[str, float]] = []
    parse_errors: list[str] = []
    parsed_sheet = ""

    if suffix in {".xlsx", ".xlsm"}:
        distribution_rows, parse_errors, parsed_sheet = _parse_with_openpyxl(file_path)
    else:
        distribution_rows, parse_errors, parsed_sheet = _parse_with_pandas_xls(file_path)

    has_distribution = has_tld_distribution(distribution_rows)

    return {
        "source_path": str(file_path.resolve()),
        "distribution_rows": distribution_rows,
        "has_parsed_distribution": has_distribution,
        "has_parsed_loads": has_distribution,
        "has_parsed_values": has_distribution,
        "parse_errors": parse_errors,
        "parsed_sheet": parsed_sheet,
    }


def parse_tld_distribution_rows(rows: list[list]) -> tuple[list[dict[str, float]], list[str]]:
    """Parse grid rows into TLD distribution entries."""
    errors: list[str] = []
    layout = _find_tld_layout(rows)
    if layout is None:
        return [], ["Could not find header row containing 'Axle group load'."]

    _header_row_index, load_column, axle_columns, data_start_row = layout

    if not axle_columns:
        return [], ["Could not find axle type columns (SAST, SADT, TAST, TADT, TRDT)."]
    if data_start_row is None or data_start_row >= len(rows):
        return [], ["Could not determine where TLD data rows begin."]

    distribution: list[dict[str, float]] = []
    for row_index, row in enumerate(rows[data_start_row:], start=data_start_row):
        if not row:
            if distribution:
                break
            continue

        load_kn = _coerce_float(_cell_at(row, load_column))
        if load_kn is None or load_kn <= 0:
            if not distribution and _row_looks_like_non_data(row, load_column, axle_columns):
                continue
            break

        entry: dict[str, float] = {"load_kn": load_kn}
        for axle_type in AXLE_TYPE_ORDER:
            percent = _coerce_float(_cell_at(row, axle_columns.get(axle_type)))
            if percent is None:
                percent = 0.0
            if percent < 0:
                errors.append(f"Negative percentage for {axle_type} at load {load_kn} kN.")
                percent = 0.0
            if percent > 100:
                errors.append(f"Percentage above 100 for {axle_type} at load {load_kn} kN.")
            entry[axle_type] = percent
        distribution.append(entry)

    if not distribution:
        return [], ["No TLD distribution rows found below the header."]
    return distribution, errors


def _find_tld_layout(
    rows: list[list],
) -> tuple[int, int, dict[str, int], int] | None:
    """Locate Austroads TLD header, axle columns, and first data row."""
    header_row_index: int | None = None
    load_column = 0

    for row_index, row in enumerate(rows[:60]):
        if not row:
            continue
        for column_index, cell in enumerate(row):
            if "axle group load" in str(cell or "").strip().lower():
                header_row_index = row_index
                load_column = column_index
                break
        if header_row_index is not None:
            break

    if header_row_index is None:
        return None

    axle_columns: dict[str, int] = {}
    axle_label_row: int | None = None
    percent_row: int | None = None

    # Prefer axle labels on rows below the header (Austroads two-row layout).
    for offset in range(1, 4):
        search_row = header_row_index + offset
        if search_row >= len(rows):
            break
        row = rows[search_row]
        if not row:
            continue
        found = _scan_row_for_axle_columns(row)
        if found:
            axle_columns = found
            axle_label_row = search_row
            break

    if not axle_columns:
        header_row = rows[header_row_index]
        found = _scan_row_for_axle_columns(header_row)
        if found:
            axle_columns = found
            axle_label_row = header_row_index

    if axle_label_row is not None:
        next_row = axle_label_row + 1
        if next_row < len(rows) and _is_percent_row(rows[next_row], axle_columns):
            percent_row = next_row

    if not axle_columns:
        return header_row_index, load_column, {}, header_row_index + 1

    if percent_row is not None:
        data_start_row = percent_row + 1
    elif axle_label_row is not None:
        data_start_row = axle_label_row + 1
    else:
        data_start_row = header_row_index + 1

    return header_row_index, load_column, axle_columns, data_start_row


def _scan_row_for_axle_columns(row: list) -> dict[str, int]:
    columns: dict[str, int] = {}
    for column_index, cell in enumerate(row):
        axle_type = _normalize_axle_label(cell)
        if axle_type:
            columns[axle_type] = column_index
    return columns


def _is_percent_row(row: list, axle_columns: dict[str, int] | None = None) -> bool:
    if axle_columns:
        cells = [_cell_at(row, column_index) for column_index in axle_columns.values()]
    else:
        cells = row
    markers = 0
    other = 0
    for cell in cells:
        text = str(cell or "").strip()
        if not text:
            continue
        if text in {"%", "percent"} or text.lower() == "percent" or "%" in text:
            markers += 1
        else:
            other += 1
    return markers > 0 and other == 0


def _row_looks_like_non_data(row: list, load_column: int, axle_columns: dict[str, int]) -> bool:
    load_text = str(_cell_at(row, load_column) or "").strip().lower()
    if load_text and not _coerce_float(_cell_at(row, load_column)):
        if any(token in load_text for token in ("axle", "load", "group", "type")):
            return True
        if "%" in load_text:
            return True
    for column_index in axle_columns.values():
        if _normalize_axle_label(_cell_at(row, column_index)):
            return True
    if _is_percent_row(row, axle_columns):
        return True
    return False


def _parse_with_openpyxl(path: Path) -> tuple[list[dict[str, float]], list[str], str]:
    try:
        from openpyxl import load_workbook
    except Exception:
        return [], ["openpyxl is required to read TLD Excel files."], ""

    try:
        workbook = load_workbook(path, data_only=True, read_only=False)
    except Exception as exc:
        return [], [f"Could not open workbook: {exc}"], ""

    try:
        for worksheet in workbook.worksheets:
            rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
            parsed_rows, errors = parse_tld_distribution_rows(rows)
            if parsed_rows:
                return parsed_rows, errors, worksheet.title
    finally:
        workbook.close()
    return [], ["No TLD distribution rows found in workbook."], ""


def _parse_with_pandas_xls(path: Path) -> tuple[list[dict[str, float]], list[str], str]:
    try:
        import pandas as pd
    except Exception:
        return [], ["pandas is required to read .xls TLD files."], ""

    try:
        sheets = pd.read_excel(path, sheet_name=None, header=None, engine="xlrd")
    except Exception as exc:
        return [], [f"Could not open workbook: {exc}"], ""

    for sheet_name, frame in sheets.items():
        rows = frame.where(frame.notna(), None).values.tolist()
        parsed_rows, errors = parse_tld_distribution_rows(rows)
        if parsed_rows:
            return parsed_rows, errors, str(sheet_name)
    return [], ["No TLD distribution rows found in workbook."], ""


def _normalize_axle_label(value: object) -> str | None:
    text = re.sub(r"\s+", "", str(value or "").strip().upper())
    if text in _TLD_AXLE_LABELS:
        return text
    return None


def _cell_at(row: list, column_index: int | None) -> object:
    if column_index is None or column_index < 0 or column_index >= len(row):
        return None
    return row[column_index]


def _coerce_float(value: object) -> float | None:
    if value is None:
        return None
    try:
        number = float(str(value).replace(",", "").strip().replace("%", ""))
    except ValueError:
        return None
    if number != number:
        return None
    return number
