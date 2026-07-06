"""Unit tests for Austroads TLD Excel parsing and weighted ESAL factors."""
from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from app.services.esal_calculator import (
    STANDARD_LOADS_KN,
    build_axle_esal_rows,
    calculate_axle_esal_per_day,
    calculate_distribution_row_factor,
    calculate_weighted_esal_factor,
    get_ldf_by_lane,
)
from app.services.traffic_esal import compute_esal_from_workbook_data
from app.services.traffic_tld_excel import parse_tld_distribution_rows, read_tld_workbook

_SAMPLE_GRID = [
    ["Axle group load (kN)", "SAST", "SADT", "TAST", "TADT", "TRDT"],
    [10, 0.2804, 3.4730, 0.0354, 0.1444, 0.0050],
    [50, 29.94, 0.0, 0.0, 0.0, 0.0],
    [80, 0.0, 0.0, 0.0, 0.0, 0.0],
    ["", "", "", "", "", ""],
]

_AUSTROADS_GRID = [
    ["Austroads Appendix TLD", None, None, None, None, None],
    [None, None, None, None, None, None],
    ["Axle group load (kN)", "Axle group type", None, None, None, None],
    [None, "SAST", "SADT", "TAST", "TADT", "TRDT"],
    [None, "%", "%", "%", "%", "%"],
    [10, 0.2804, 3.4730, 0.0354, 0.1444, 0.0050],
    [20, 7.8270, 8.6960, 0.2377, 0.5755, 0.1568],
    [50, 29.94, 0.0, 0.0, 0.0, 0.0],
    ["", "", "", "", "", ""],
]

_EXAMPLE_WORKBOOK = {
    "summary_total_row": [""] + [297, 0, 0, 0, 0, 0, 0, 0, 562, 5950, 0, 8687, 0, 0, 0, 0, 0, 0],
}


class TldDistributionTests(unittest.TestCase):
    def test_parser_reads_single_row_header(self) -> None:
        rows, errors = parse_tld_distribution_rows(_SAMPLE_GRID)
        self.assertEqual(errors, [])
        self.assertEqual(len(rows), 3)
        self.assertEqual(rows[0]["load_kn"], 10.0)
        self.assertAlmostEqual(rows[0]["SAST"], 0.2804)

    def test_parser_skips_label_row_when_data_start_is_early(self) -> None:
        grid = [
            ["Austroads Appendix TLD", None, None, None, None, None],
            [None, None, None, None, None, None],
            [None, None, None, None, None, None],
            ["Axle group load", "SAST", "SADT", "TAST", "TADT", "TRDT"],
            [None, "SAST", "SADT", "TAST", "TADT", "TRDT"],
            [None, "%", "%", "%", "%", "%"],
            [10, 0.2804, 3.4730, 0.0354, 0.1444, 0.0050],
            [50, 29.94, 0.0, 0.0, 0.0, 0.0],
        ]
        rows, errors = parse_tld_distribution_rows(grid)
        self.assertEqual(errors, [])
        self.assertGreater(len(rows), 0)
        self.assertEqual(rows[0]["load_kn"], 10.0)

    def test_parser_reads_austroads_two_row_header(self) -> None:
        rows, errors = parse_tld_distribution_rows(_AUSTROADS_GRID)
        self.assertEqual(errors, [])
        self.assertGreater(len(rows), 0)
        self.assertEqual(rows[0]["load_kn"], 10.0)
        self.assertAlmostEqual(rows[0]["SAST"], 0.2804)
        self.assertAlmostEqual(rows[0]["SADT"], 3.4730)
        self.assertAlmostEqual(rows[1]["load_kn"], 20.0)
        self.assertAlmostEqual(rows[1]["SAST"], 7.8270)
        self.assertAlmostEqual(rows[2]["load_kn"], 50.0)
        self.assertAlmostEqual(rows[2]["SAST"], 29.94)

    def test_weighted_factor_uses_percent_over_100(self) -> None:
        rows, _errors = parse_tld_distribution_rows(_AUSTROADS_GRID)
        factor = calculate_weighted_esal_factor("SAST", rows)
        self.assertGreater(factor, 0.0)
        expected = 0.0
        for row in rows:
            expected += calculate_distribution_row_factor(
                row["load_kn"],
                STANDARD_LOADS_KN["SAST"],
                row["SAST"],
            )
        self.assertAlmostEqual(factor, expected)

    def test_use_tld_esal_per_day_equals_count_times_weighted_factor_times_ldf(self) -> None:
        rows, _errors = parse_tld_distribution_rows(_AUSTROADS_GRID)
        counts = {"steering_sast": 100, "tast": 0, "sadt": 0, "tadt": 0, "trdt": 0}
        axle_rows = build_axle_esal_rows(
            counts,
            tld_distribution=rows,
            use_tld=True,
            tld_loads_ready=True,
            lane_each_direction=2,
        )
        sast_row = axle_rows[0]
        weighted = calculate_weighted_esal_factor("SAST", rows)
        expected = calculate_axle_esal_per_day(100, weighted, get_ldf_by_lane(2))
        self.assertAlmostEqual(sast_row.esal_per_day, expected)
        self.assertIsNone(sast_row.actual_load_kn)

    def test_assume_standard_load_still_uses_factor_one(self) -> None:
        counts = {"steering_sast": 100, "tast": 0, "sadt": 0, "tadt": 0, "trdt": 0}
        axle_rows = build_axle_esal_rows(counts, use_tld=False, lane_each_direction=2)
        self.assertEqual(axle_rows[0].esal_factor, 1.0)
        self.assertAlmostEqual(axle_rows[0].esal_per_day, 90.0)

    def test_use_tld_workbook_integration(self) -> None:
        distribution_rows, _errors = parse_tld_distribution_rows(_AUSTROADS_GRID)
        tld_data = {
            "distribution_rows": distribution_rows,
            "has_parsed_distribution": True,
            "has_parsed_loads": True,
        }
        self.assertGreater(len(distribution_rows), 0)
        self.assertTrue(tld_data["has_parsed_loads"])

        result = compute_esal_from_workbook_data(
            _EXAMPLE_WORKBOOK,
            use_tld=True,
            tld_data=tld_data,
            lane_count=1,
            growth_rate=0.03,
            pavement_design_years=25,
        )
        self.assertTrue(result.tld_loads_ready)
        self.assertIsNone(result.tld_message)
        self.assertGreater(result.total_esal_per_day, 0)
        self.assertGreater(result.table_periods[0].total_esal, 0)
        self.assertNotEqual(result.table_periods[0].traffic_class, "—")

    def test_parser_reads_kiec_merged_header_layout(self) -> None:
        grid = [
            ["Austroads Appendix TLD", None, None, None, None, None],
            [None, None, None, None, None, None],
            [None, "Axle group type", None, None, None, None],
            ["Axle group load (kN)", "SAST", "SADT", "TAST", "TADT", "TRDT"],
            [None, "%", "%", "%", "%", "%"],
            [10, 0.2804, 3.4730, 0.0354, 0.1444, 0.0050],
        ]
        rows, errors = parse_tld_distribution_rows(grid)
        self.assertEqual(errors, [])
        self.assertEqual(rows[0]["load_kn"], 10.0)
        self.assertAlmostEqual(rows[0]["SADT"], 3.4730)

    def test_read_kiec_workbook_with_merged_cells(self) -> None:
        try:
            from openpyxl import Workbook
            from openpyxl.utils import get_column_letter
        except ImportError:
            self.skipTest("openpyxl not available")

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "TLD"
        sheet["A1"] = "Austroads Appendix TLD"
        sheet.merge_cells("A3:A5")
        sheet["A3"] = "Axle group load (kN)"
        sheet.merge_cells("B3:F3")
        sheet["B3"] = "Axle group type"
        for column, label in enumerate(["SAST", "SADT", "TAST", "TADT", "TRDT"], start=2):
            sheet.cell(row=4, column=column, value=label)
            sheet.cell(row=5, column=column, value="%")
        data_rows = [
            (10, 0.2804, 3.4730, 0.0354, 0.1444, 0.0050),
            (20, 7.8270, 8.6960, 0.2377, 0.5755, 0.1568),
        ]
        for row_index, values in enumerate(data_rows, start=6):
            for column_index, value in enumerate(values, start=1):
                sheet.cell(row=row_index, column=column_index, value=value)

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as handle:
            temp_path = handle.name
        try:
            workbook.save(temp_path)
            workbook.close()
            tld_data = read_tld_workbook(temp_path)
        finally:
            Path(temp_path).unlink(missing_ok=True)

        self.assertTrue(tld_data["has_parsed_loads"])
        self.assertGreaterEqual(len(tld_data["distribution_rows"]), 2)
        self.assertEqual(tld_data["distribution_rows"][0]["load_kn"], 10.0)

        try:
            from openpyxl import Workbook
        except ImportError:
            self.skipTest("openpyxl not available")

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "TLD"
        for row_index, row in enumerate(_AUSTROADS_GRID, start=1):
            for column_index, value in enumerate(row, start=1):
                if value is not None:
                    sheet.cell(row=row_index, column=column_index, value=value)

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as handle:
            temp_path = handle.name
        try:
            workbook.save(temp_path)
            workbook.close()
            tld_data = read_tld_workbook(temp_path)
        finally:
            Path(temp_path).unlink(missing_ok=True)

        self.assertTrue(tld_data["has_parsed_loads"])
        self.assertGreater(len(tld_data["distribution_rows"]), 0)
        weighted = calculate_weighted_esal_factor("SAST", tld_data["distribution_rows"])
        self.assertGreater(weighted, 0.0)


if __name__ == "__main__":
    unittest.main()
